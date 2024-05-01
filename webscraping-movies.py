import ssl 
ssl._create_default_https_context = ssl._create_unverified_context
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'
page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)
##
wb = xl.Workbook()
ws =wb.active
ws.title = 'Box Office Report'
##
ws.append(['No.','Movie Title','Release Data','Total Gross','Theaters','Average per theater'])
for row in soup.findAll('tr')[1:6]:
    td = row.findAll('td')
    total = int(td[7].text.replace('$','').replace(',',''))
    theater = int(td[6].text.replace('$','').replace(',',''))
    per = total/theater
    ws.append([td[0].text,td[1].text,td[8].text,td[7].text,td[6].text,f'${per:,.2f}'])
##
##
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 28

ws.cell(1,1).font = Font(size=16,bold=True)
ws.cell(1,2).font = Font(size=16,bold=True)
ws.cell(1,3).font = Font(size=16,bold=True)
ws.cell(1,4).font = Font(size=16,bold=True)
ws.cell(1,5).font = Font(size=16,bold=True)
ws.cell(1,6).font = Font(size=16,bold=True)

wb.save('BoxOfficeReport.xlsx')